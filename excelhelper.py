import pandas as pd
from openpyxl import load_workbook

def copy_data_between_excel(data_file, blank_file, data_location):
    """
    Copies data from one Excel file to another based on the provided data location.
    
    Parameters:
    - data_file: str, path to the Excel file with data.
    - blank_file: str, path to the blank Excel file.
    - data_location: dict, keys are populated and blank sheet names, and values are dictionaries with source and target locations. 
                     Example:
                     {
                         ('Sheet1', 'BlankSheet1'): {'source': 'A1:D10', 'target': 'A1'},
                         ('Sheet2', 'BlankSheet2'): {'source': 'B5:E15', 'target': 'C5'}
                     }
    """
    
    # Load both Excel files
    data_wb = load_workbook(data_file)
    blank_wb = load_workbook(blank_file)
    
    for sheet_pair, location in data_location.items():
        populated_sheet_name, blank_sheet_name = sheet_pair
        source_range = location['source']
        target_cell = location['target']
        
        # Get sheets from data and blank workbooks
        data_ws = data_wb[populated_sheet_name]
        blank_ws = blank_wb[blank_sheet_name]
        
        # Read the source data from the data file
        data_df = pd.read_excel(data_file, sheet_name=populated_sheet_name, engine='openpyxl')
        
        # Extract the source data
        source_data = data_ws[source_range]
        
        # Find the starting row and column for the target cell
        target_row = int(target_cell[1:])  # Get the row number from target_cell, e.g., A1 -> 1
        target_col = target_cell[0]        # Get the column letter from target_cell, e.g., A1 -> A
        
        # Copy the data into the blank worksheet
        for i, row in enumerate(source_data, start=target_row):
            for j, cell in enumerate(row, start=ord(target_col) - ord('A') + 1):  # Convert column letter to index
                blank_ws.cell(row=i, column=j, value=cell.value)
        
    # Save the modified blank workbook
    blank_wb.save(blank_file)
    print(f"Data copied successfully to {blank_file}")

# Define the sheet names and data ranges
data_location = {
    ('Sheet1', 'BlankSheet1'): {'source': 'A1:D10', 'target': 'A1'},
    ('Sheet2', 'BlankSheet2'): {'source': 'B5:E15', 'target': 'C5'}
}

# Call the function to copy data
copy_data_between_excel('data.xlsx', 'blank.xlsx', data_location)

